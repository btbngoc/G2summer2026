import json
import os
import re
from openpyxl import load_workbook


BASE_DIR = os.path.dirname(__file__)
EXCEL_CANDIDATES = [
    os.path.join(BASE_DIR, 'Danh sách phòng.xlsx'),
    os.path.join(BASE_DIR, 'assets', 'Danh sách phòng.xlsx'),
]
OUT_JSON = os.path.join(BASE_DIR, 'assets', 'rooms.json')


def clean_text(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def parse_members(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [clean_text(v) for v in value if clean_text(v)]
    text = clean_text(value)
    if not text:
        return []
    # Split by common delimiters: ; / , |
    parts = re.split(r'[;|,/]+', text)
    return [p.strip() for p in parts if p.strip()]


def normalize_room_row(row):
    # Support common Excel layouts:
    # LDAP | Xe | Số phòng | Members |Ghi chú
    # or the first 4 columns in order.
    ldap = clean_text(row[0]) if len(row) > 0 else ''
    bus = clean_text(row[1]) if len(row) > 1 else ''
    room = clean_text(row[2]) if len(row) > 2 else ''
    members_raw = clean_text(row[3]) if len(row) > 3 else ''
    note = clean_text(row[4]) if len(row) > 4 else ''

    # If the workbook uses header labels, skip them.
    #if ldap.lower() in {'ldap', 'mã ldap', 'mã'} and name.lower() in {'họ tên', 'tên', 'name'}:
        #return None

    members = parse_members(members_raw)
    if not members and len(row) >= 5:
        # If the fourth cell contains the member list and the fifth contains note,
        # use the list from the 4th column for member names and 5th as note.
        pass

    # If the row only has a name and room but no LDAP, ignore it.
    if not ldap and not bus and not room:
        return None

    return {
        'ldap': ldap,
        'bus': bus,
        'room': room,
        'members': members if members else [ldap],
        'note': note,
    }


def read_from_excel(path):
    wb = load_workbook(path, data_only=True)
    rooms = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            # Skip totally empty rows.
            if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                continue
            room = normalize_room_row(row)
            if room:
                rooms.append(room)
    return rooms


def main():
    excel_path = None
    for candidate in EXCEL_CANDIDATES:
        if os.path.exists(candidate):
            excel_path = candidate
            break

    if not excel_path:
        print('Không tìm thấy file Danh sách phòng.xlsx. Hãy đặt file vào thư mục gốc hoặc assets/.')
        print('Script sẽ tạo file rooms.json rỗng để frontend vẫn chạy.')
        os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
        with open(OUT_JSON, 'w', encoding='utf-8') as f:
            json.dump([], f, ensure_ascii=False, indent=2)
        return

    rooms = read_from_excel(excel_path)
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(rooms, f, ensure_ascii=False, indent=2)

    print(f'Đã tạo {OUT_JSON} với {len(rooms)} phòng từ {excel_path}')


if __name__ == '__main__':
    main()
