import json, sys
from openpyxl import load_workbook
from datetime import time, datetime
import os
import re

def safe(val):
    if isinstance(val, (time, datetime)):
        return val.isoformat()
    return val

path = r"C:\\Users\\btbngoc\\Downloads\\summer2026-20260611T045412Z-3-001\\summer2026\\assets\\Hạ Long 3N2D updated 18-6_7301.xlsx"

wb = load_workbook(path, data_only=True)
out = {}
for ws in wb.worksheets:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([safe(cell) for cell in row])
    out[ws.title] = rows
with open('itinerary.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

# Also create a web-friendly itinerary JSON used by the front-end
try:
    sheet_name = None
    for k in out.keys():
        if 'lịch trình' in k.lower() or 'lich trinh' in k.lower() or 'lịch trình chi tiết' in k.lower():
            sheet_name = k
            break
    if sheet_name is None:
        sheet_name = next(iter(out.keys()))
    rows = out.get(sheet_name, [])

    web_itinerary = []
    current = None
    time_re = re.compile(r'[0-2]?\d[:h]?[0-5]?\d|\d{1,2}:\d{2}[-–]\d{1,2}:\d{2}')
    for row in rows:
        if not row:
            continue
        c0 = '' if row[0] is None else str(row[0]).strip()
        c1 = '' if len(row) < 2 or row[1] is None else str(row[1]).strip()
        if not c0 and not c1:
            continue
        up0 = c0.upper()
        if up0.startswith('NGÀY') or up0.startswith('NGAY'):
            if current:
                web_itinerary.append(current)
            current = {'day': c0, 'sub': c1 or '', 'items': []}
            continue
        if c0 and (':' in c0 or '-' in c0 or time_re.search(c0)):
            if current is None:
                continue
            item_time = c0.replace('.000000', '').replace('00:00:00', '').strip()
            title = c1 or ''
            desc_parts = [title]
            if len(row) > 2:
                for extra in row[2:5]:
                    if extra:
                        desc_parts.append(str(extra).strip())
            desc = ' '.join([p for p in desc_parts if p])
            current['items'].append({'time': item_time, 't': title, 'ic': '', 'd': desc, 'loc': '', 'note': ''})
            continue
    if current:
        web_itinerary.append(current)

    assets_dir = os.path.join(os.path.dirname(__file__), 'assets')
    os.makedirs(assets_dir, exist_ok=True)
    web_path = os.path.join(assets_dir, 'itinerary_web.json')
    with open(web_path, 'w', encoding='utf-8') as wf:
        json.dump(web_itinerary, wf, ensure_ascii=False, indent=2)
    print('Wrote web-friendly itinerary to', web_path)
except Exception as e:
    print('Failed to write web itinerary:', e)

